import hashlib
import io
import ipaddress
import json
import os
import shutil
import subprocess
import tempfile
import time
import uuid
from datetime import date, datetime, timedelta

import pyzipper
import qrcode
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv
from flask import (
    Flask,
    Response,
    abort,
    flash,
    has_request_context,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFError, CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.exc import IntegrityError
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

load_dotenv()

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
RUNTIME_DIR = os.path.join(BASE_DIR, "runtime")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
BACKUP_FOLDER = os.getenv("BACKUP_FOLDER") or os.path.join(RUNTIME_DIR, "backups")


def bool_env(name, default="0"):
    return os.getenv(name, default).strip().lower() in {"1", "true", "yes", "on"}


def csv_env(name, default=""):
    return [item.strip() for item in os.getenv(name, default).split(",") if item.strip()]


def utcnow():
    return datetime.utcnow()


def as_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            parsed = datetime.strptime(text[:19], fmt)
            return parsed.date()
        except ValueError:
            continue
    return None


def parse_datetime_local(value):
    if value in (None, ""):
        return None
    for fmt in ("%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


APP_NAME = "Γραφείο Ασφαλείας 251 ΓΝΑ"

ROLES = {
    "admin": {
        "label": "Admin",
        "description": "Χρήστες, backup, audit και πλήρης λειτουργία εφαρμογής.",
    },
    "security_office": {
        "label": "Γραφείο Ασφαλείας",
        "description": "Πλήρης επιχειρησιακή χρήση χωρίς χρήστες και backup.",
    },
    "transit_office": {
        "label": "Γραφείο Διακίνησης",
        "description": "Ημερήσια κατάσταση, διακίνηση και ασθενείς χωρίς διοικητικό εξιτήριο.",
    },
    "security_point": {
        "label": "Security point",
        "description": "Έλεγχος πύλης, QR scan, νοσηλευόμενοι και επισκέπτες.",
    },
}

PATIENT_COLUMNS = [
    ("pat", "ΑΡ.ΠΕΡ."),
    ("vip", "ΒΑΘΜΟΣ"),
    ("pcd", "ΕΙΔΙΚ"),
    ("chr", "ΟΝΟΜΑ"),
    ("name", "ΕΠΙΘΕΤΟ"),
    ("unit_name", "ΜΟΝΑΔΑ"),
    ("street", "ΔΙΕΥΘΥΝΣΗ"),
    ("city", "ΠΟΛΗ"),
    ("cfe", "ΙΔΙΟΤΗΤΑ"),
    ("cfe_type01", "ΣΤΡ.ΠΡΣ-ΛΟΙΠΟΙ"),
    ("hm_eis", "ΗΜ.ΕΙΣ."),
    ("wds", "ΟΡΟΦΟΣ"),
    ("dep", "ΚΛΙΝΙΚΗ"),
    ("room", "ΔΩΜΑΤΙΟ"),
]

PATIENT_MEDICO_MAP = {
    "VIP": "vip",
    "PCD": "pcd",
    "CHR": "chr",
    "NAME": "name",
    "UNIT_NAME": "unit_name",
    "STREET": "street",
    "CITY": "city",
    "CFE": "cfe",
    "CFE_TYPE01": "cfe_type01",
    "HM_EIS": "hm_eis",
    "WDS": "wds",
    "DEP": "dep",
    "ROOM": "room",
    "PAT": "pat",
}

MOVEMENT_COLUMNS = [
    ("PAT", "ΥΓ.ΚΩΔ."),
    ("VIP", "ΒΑΘΜΟΣ"),
    ("PCD", "ΕΙΔΙΚ"),
    ("CHR", "ΟΝΟΜΑ"),
    ("NAME", "ΕΠΙΘΕΤΟ"),
    ("PERSCODE", "ΑΡ.ΤΑΥΤ."),
    ("UNIT_NAME", "ΜΟΝΑΔΑ"),
    ("STREET", "ΔΙΕΥΘΥΝΣΗ"),
    ("ZIP", "Τ.Κ."),
    ("CITY", "ΠΟΛΗ"),
    ("PHONE", "ΤΗΛ."),
    ("CFE", "ΙΔΙΟΤΗΤΑ"),
    ("CFE_TYPE01", "ΣΤΡ.ΠΡΣ-ΛΟΙΠΟΙ"),
    ("HM_EIS", "ΗΜΕΡ.ΕΙΣ."),
    ("AMKA_DOC", "ΑΜΚΑ ΙΑΤΡΟΥ"),
    ("WDS", "ΟΡΟΦΟΣ"),
    ("DEP", "ΚΛΙΝΙΚΗ"),
    ("ROOM", "ΔΩΜΑΤΙΟ"),
]

NO_EXIT_COLUMNS = [
    ("PAT", "ΑΡ.ΠΕΡ."),
    ("NAMECHR", "ΟΝΟΜ/ΜΟ"),
    ("DISDCALC", "ΗΜ/ΝΙΑ"),
    ("DEP", "ΚΛΙΝΙΚΗ"),
    ("NAME", "ΤΑΜΕΙΟ"),
    ("VIP", "ΒΑΘΜΟΣ"),
    ("KCODE", "KCODE"),
]

PATIENT_QUERY = """
SELECT "VIP", "PCD", "CHR", "NAME", "UNIT_NAME", "STREET", "CITY", "CFE",
       "CFE_TYPE01", "HM_EIS", "WDS", "DEP", "ROOM", "PAT"
FROM "HMERKATNOSIL"
"""

MOVEMENT_QUERY = """
SELECT "PAT", "VIP", "PCD", "CHR", "NAME", "PERSCODE", "UNIT_NAME", "STREET",
       "ZIP", "CITY", "PHONE", "CFE", "CFE_TYPE01", "HM_EIS", "AMKA_DOC",
       "WDS", "DEP", "ROOM"
FROM "HMERKATNOSIL"
"""

NO_EXIT_QUERY = """
SELECT DISTINCT X1100PAT.PAT, X1100PAT.NAMECHR, X1100PAT.DISDCALC, X1280DIA.DEP,
       X8001DEB.NAME, X1000PER.VIP, KEN_KDATA.KCODE
FROM X1000PER, X1100PAT, X1280DIA, X1150COG, X8001DEB, KEN_KDATA
WHERE (X1000PER.PER = X1100PAT.PER)
  AND (X1280DIA.PAT = X1100PAT.PAT)
  AND (X1150COG.PAT = X1100PAT.PAT)
  AND (X8001DEB.DEB = X1150COG.DEB)
  AND (X1100PAT.PAT = KEN_KDATA.EPISODE(+))
  AND (X1100PAT.DISD = '31-12-2099')
  AND (X1100PAT.TYP = 'S')
  AND (X1100PAT.DISDCALC IS NOT NULL)
  AND (X1280DIA.DIT = 'ΕΞΙ')
ORDER BY X1100PAT.NAMECHR, X1100PAT.DISDCALC
"""

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1, x_prefix=1)
app.secret_key = os.getenv("FLASK_SECRET_KEY") or os.getenv("SECRET_KEY") or "dev-only-change-me"
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=int(os.getenv("SESSION_HOURS", "8")))
app.config["SESSION_COOKIE_NAME"] = os.getenv("SESSION_COOKIE_NAME", "sec_office_session")
app.config["SESSION_COOKIE_SECURE"] = bool_env("SESSION_COOKIE_SECURE", "0")
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("DATABASE_URL") or "sqlite:///" + os.path.join(RUNTIME_DIR, "sec_office.sqlite3")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = int(os.getenv("MAX_UPLOAD_MB", "25")) * 1024 * 1024

os.makedirs(RUNTIME_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

csrf = CSRFProtect(app)
db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Παρακαλώ συνδεθείτε για να συνεχίσετε."
login_manager.login_message_category = "warning"
login_manager.session_protection = "basic"

TRUST_SSO_HEADERS = bool_env("TRUST_SSO_HEADERS", "0")
SSO_TRUSTED_PROXY_CIDRS = csv_env("SSO_TRUSTED_PROXY_CIDRS", "127.0.0.1/32,172.16.0.0/12")
SSO_APP_USER_GROUP = os.getenv("SSO_APP_USER_GROUP", "/apps/sec-office/users")
SSO_APP_ADMIN_GROUP = os.getenv("SSO_APP_ADMIN_GROUP", "/apps/sec-office/admins")
SSO_GLOBAL_ADMIN_GROUP = os.getenv("SSO_GLOBAL_ADMIN_GROUP", "/apps/global/admins")
SSO_SECURITY_OFFICE_GROUP = os.getenv("SSO_SECURITY_OFFICE_GROUP", "/apps/sec-office/security-office")
SSO_TRANSIT_OFFICE_GROUP = os.getenv("SSO_TRANSIT_OFFICE_GROUP", "/apps/sec-office/transit-office")
SSO_SECURITY_POINT_GROUP = os.getenv("SSO_SECURITY_POINT_GROUP", "/apps/sec-office/security-point")
CENTRAL_AUTH_REALM = os.getenv("CENTRAL_AUTH_REALM", "intranet")
CENTRAL_AUTH_ADMIN_URL = os.getenv("CENTRAL_AUTH_ADMIN_URL", "https://auth.251gh.local/admin/")
CENTRAL_AUTH_USERS_URL = os.getenv(
    "CENTRAL_AUTH_USERS_URL",
    f"https://auth.251gh.local/admin/master/console/#/{CENTRAL_AUTH_REALM}/users",
)
CENTRAL_AUTH_GROUPS_URL = os.getenv(
    "CENTRAL_AUTH_GROUPS_URL",
    f"https://auth.251gh.local/admin/master/console/#/{CENTRAL_AUTH_REALM}/groups",
)

MEDICO_SAMPLE_MODE = bool_env("MEDICO_SAMPLE_MODE", "0")
MEDICO_SYNC_INTERVAL_MINUTES = int(os.getenv("MEDICO_SYNC_INTERVAL_MINUTES", "15"))
MEDICO_DELETE_MISSING = bool_env("MEDICO_DELETE_MISSING", "1")
MEDICO_SKIP_DELETE_ON_EMPTY = bool_env("MEDICO_SKIP_DELETE_ON_EMPTY", "1")
ORACLE_CLIENT_LIB_DIR = os.getenv("ORACLE_CLIENT_LIB_DIR", "").strip()
ORACLE_CONFIG_DIR = os.getenv("ORACLE_CONFIG_DIR", "").strip()
BACKUP_RETENTION_COUNT = int(os.getenv("BACKUP_RETENTION_COUNT", "30"))
BACKUP_INCLUDE_DATABASE = bool_env("BACKUP_INCLUDE_DATABASE", "1")

_oracle_client_initialized = False


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(40), default="security_point", nullable=False)
    is_active_local = db.Column(db.Boolean, default=True, nullable=False)
    created_at = db.Column(db.DateTime, default=utcnow, nullable=False)

    @property
    def is_admin(self):
        return self.role == "admin"

    @property
    def auth_method(self):
        return "local"


class SSOUser(UserMixin):
    def __init__(self, username, email=None, groups=None, role="security_point"):
        self.id = f"sso:{username}"
        self.username = username
        self.email = email
        self.sso_groups = groups or []
        self.role = role

    @property
    def is_admin(self):
        return self.role == "admin"

    @property
    def auth_method(self):
        return "sso"

    @property
    def is_active(self):
        return True


class PatientRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    qr_token = db.Column(db.String(36), unique=True, nullable=False, default=lambda: str(uuid.uuid4()))
    pat = db.Column(db.String(80), unique=True, nullable=False, index=True)
    vip = db.Column(db.String(100))
    pcd = db.Column(db.String(100))
    chr = db.Column(db.String(255))
    name = db.Column(db.String(255))
    unit_name = db.Column(db.String(255))
    street = db.Column(db.String(255))
    city = db.Column(db.String(255))
    cfe = db.Column(db.String(255))
    cfe_type01 = db.Column(db.String(255))
    hm_eis = db.Column(db.Date)
    wds = db.Column(db.String(100))
    dep = db.Column(db.String(255), index=True)
    room = db.Column(db.String(100), index=True)
    escort_one = db.Column(db.String(255))
    escort_two = db.Column(db.String(255))
    security_notes = db.Column(db.Text)
    raw_medico_json = db.Column(db.Text)
    medico_hash = db.Column(db.String(64))
    first_seen_at = db.Column(db.DateTime, default=utcnow, nullable=False)
    last_seen_at = db.Column(db.DateTime, default=utcnow, nullable=False)
    last_synced_at = db.Column(db.DateTime)
    last_scanned_at = db.Column(db.DateTime)
    last_scanned_by = db.Column(db.String(120))
    updated_at = db.Column(db.DateTime, default=utcnow, onupdate=utcnow, nullable=False)

    @property
    def full_name(self):
        return " ".join(part for part in [self.chr, self.name] if part).strip() or "-"


class VisitorRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    qr_token = db.Column(db.String(36), unique=True, nullable=False, default=lambda: str(uuid.uuid4()))
    full_name = db.Column(db.String(255), nullable=False, index=True)
    identity_number = db.Column(db.String(120), index=True)
    phone = db.Column(db.String(80))
    organization = db.Column(db.String(255))
    role_title = db.Column(db.String(255))
    visit_reason = db.Column(db.String(255))
    destination = db.Column(db.String(255), index=True)
    access_scope = db.Column(db.String(255))
    authorization_doc_no = db.Column(db.String(120))
    authorization_issuer = db.Column(db.String(255))
    authorization_date = db.Column(db.Date)
    valid_from = db.Column(db.DateTime)
    valid_until = db.Column(db.DateTime)
    status = db.Column(db.String(40), default="active", nullable=False, index=True)
    notes = db.Column(db.Text)
    authorization_filename = db.Column(db.String(255))
    authorization_original_filename = db.Column(db.String(255))
    created_by = db.Column(db.String(120))
    updated_by = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=utcnow, onupdate=utcnow, nullable=False)
    last_scanned_at = db.Column(db.DateTime)
    last_scanned_by = db.Column(db.String(120))

    @property
    def is_currently_valid(self):
        now = utcnow()
        if self.status != "active":
            return False
        if self.valid_from and now < self.valid_from:
            return False
        if self.valid_until and now > self.valid_until:
            return False
        return True


class SyncLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    started_at = db.Column(db.DateTime, default=utcnow, nullable=False)
    ended_at = db.Column(db.DateTime)
    status = db.Column(db.String(40), default="RUNNING", nullable=False)
    rows_seen = db.Column(db.Integer, default=0, nullable=False)
    inserted = db.Column(db.Integer, default=0, nullable=False)
    updated = db.Column(db.Integer, default=0, nullable=False)
    deleted = db.Column(db.Integer, default=0, nullable=False)
    message = db.Column(db.Text)
    actor = db.Column(db.String(120))


class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=utcnow, nullable=False, index=True)
    actor = db.Column(db.String(120), index=True)
    role = db.Column(db.String(40), index=True)
    auth_method = db.Column(db.String(20))
    action = db.Column(db.String(80), nullable=False, index=True)
    target_type = db.Column(db.String(80))
    target_id = db.Column(db.String(120))
    details = db.Column(db.Text)
    ip_address = db.Column(db.String(80))


class BackupRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=utcnow, nullable=False)
    created_by = db.Column(db.String(120))
    auth_method = db.Column(db.String(20))
    backup_type = db.Column(db.String(40), default="manual", nullable=False)
    filename = db.Column(db.String(255), nullable=False)
    path = db.Column(db.Text, nullable=False)
    size_bytes = db.Column(db.Integer, default=0, nullable=False)
    sha256 = db.Column(db.String(64))
    status = db.Column(db.String(40), default="CREATED", nullable=False)
    verify_message = db.Column(db.Text)
    verified_at = db.Column(db.DateTime)


@login_manager.user_loader
def load_user(user_id):
    if user_id and user_id.startswith("sso:"):
        username = session.get("sso_username") or user_id[4:]
        return SSOUser(
            username=username,
            email=session.get("sso_email"),
            groups=session.get("sso_groups", []),
            role=session.get("sso_role", "security_point"),
        )
    try:
        return User.query.get(int(user_id))
    except (TypeError, ValueError):
        return None


def current_actor():
    if has_request_context() and current_user.is_authenticated:
        return current_user.username
    return "anonymous" if has_request_context() else "system"


def current_role():
    if has_request_context() and current_user.is_authenticated:
        return getattr(current_user, "role", "")
    return ""


def role_label(role):
    return ROLES.get(role, {}).get("label", role or "-")


def log_action(action, target_type=None, target_id=None, details=None):
    actor = current_actor()
    role = current_role()
    auth_method = getattr(current_user, "auth_method", None) if has_request_context() and current_user.is_authenticated else None
    entry = AuditLog(
        actor=actor,
        role=role,
        auth_method=auth_method,
        action=action,
        target_type=target_type,
        target_id=as_text(target_id)[:120] if target_id is not None else None,
        details=details,
        ip_address=request.headers.get("X-Forwarded-For", request.remote_addr) if has_request_context() else None,
    )
    db.session.add(entry)


def parse_sso_groups(raw):
    if not raw:
        return []
    groups = []
    for chunk in raw.replace(";", ",").split(","):
        item = chunk.strip()
        if item:
            groups.append(item)
    return groups


def is_trusted_sso_proxy():
    remote = request.remote_addr or ""
    try:
        remote_ip = ipaddress.ip_address(remote)
    except ValueError:
        return False
    for cidr in SSO_TRUSTED_PROXY_CIDRS:
        try:
            if remote_ip in ipaddress.ip_network(cidr, strict=False):
                return True
        except ValueError:
            continue
    return False


def sso_role_from_groups(groups):
    if SSO_APP_ADMIN_GROUP in groups or SSO_GLOBAL_ADMIN_GROUP in groups:
        return "admin"
    if SSO_SECURITY_OFFICE_GROUP in groups:
        return "security_office"
    if SSO_TRANSIT_OFFICE_GROUP in groups:
        return "transit_office"
    if SSO_SECURITY_POINT_GROUP in groups or SSO_APP_USER_GROUP in groups:
        return "security_point"
    return None


def get_sso_user_from_headers():
    if not TRUST_SSO_HEADERS or not is_trusted_sso_proxy():
        return None
    groups = parse_sso_groups(request.headers.get("X-SSO-Groups", ""))
    role = sso_role_from_groups(groups)
    if not role:
        return None
    username = (
        request.headers.get("X-SSO-Preferred-Username")
        or request.headers.get("X-SSO-User")
        or request.headers.get("X-SSO-Email")
        or ""
    ).strip()
    if not username:
        return None
    return SSOUser(
        username=username,
        email=(request.headers.get("X-SSO-Email") or "").strip() or None,
        groups=groups,
        role=role,
    )


@app.before_request
def sync_sso_session():
    if request.endpoint in {"static", "health", "robots"}:
        return None
    sso_user = get_sso_user_from_headers()
    if not sso_user:
        return None
    if current_user.is_authenticated and getattr(current_user, "auth_method", None) == "sso":
        if current_user.username == sso_user.username and current_user.role == sso_user.role:
            return None
    session["sso_username"] = sso_user.username
    session["sso_email"] = sso_user.email
    session["sso_groups"] = sso_user.sso_groups
    session["sso_role"] = sso_user.role
    login_user(sso_user)
    log_action("LOGIN", "SSO", sso_user.username, f"SSO login as {role_label(sso_user.role)}")
    db.session.commit()
    return None


def has_role(*roles):
    return current_user.is_authenticated and getattr(current_user, "role", None) in roles


def is_admin_user():
    return has_role("admin")


def can_manage_users():
    return has_role("admin")


def can_manage_backups():
    return has_role("admin")


def can_view_audit():
    return has_role("admin", "security_office")


def can_view_patients():
    return has_role("admin", "security_office", "security_point", "transit_office")


def can_edit_patient_access():
    return has_role("admin", "security_office", "security_point")


def can_sync_patients():
    return has_role("admin", "security_office")


def can_view_visitors():
    return has_role("admin", "security_office", "security_point")


def can_manage_visitors():
    return has_role("admin", "security_office")


def can_scan_qr():
    return has_role("admin", "security_office", "security_point")


def can_view_transit():
    return has_role("admin", "security_office", "transit_office")


def require_permission(check):
    if not check():
        abort(403)


@app.context_processor
def inject_helpers():
    return {
        "app_name": APP_NAME,
        "roles": ROLES,
        "role_label": role_label,
        "is_admin_user": is_admin_user,
        "can_manage_users": can_manage_users,
        "can_manage_backups": can_manage_backups,
        "can_view_audit": can_view_audit,
        "can_view_patients": can_view_patients,
        "can_edit_patient_access": can_edit_patient_access,
        "can_sync_patients": can_sync_patients,
        "can_view_visitors": can_view_visitors,
        "can_manage_visitors": can_manage_visitors,
        "can_scan_qr": can_scan_qr,
        "can_view_transit": can_view_transit,
        "patient_columns": PATIENT_COLUMNS,
    }


@app.template_filter("datetime_el")
def datetime_el(value):
    if not value:
        return "-"
    if isinstance(value, str):
        return value
    return value.strftime("%d/%m/%Y %H:%M")


@app.template_filter("date_el")
def date_el(value):
    if not value:
        return "-"
    if isinstance(value, str):
        return value
    return value.strftime("%d/%m/%Y")


@app.template_filter("filesize")
def filesize(value):
    try:
        size = int(value or 0)
    except (TypeError, ValueError):
        return "-"
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024:
            return f"{size:.0f} {unit}"
        size /= 1024
    return f"{size:.1f} TB"


def normalize_external_row(row):
    normalized = {}
    for key, value in row.items():
        upper = str(key).upper()
        if isinstance(value, datetime):
            normalized[upper] = value.date().isoformat()
        elif isinstance(value, date):
            normalized[upper] = value.isoformat()
        else:
            normalized[upper] = as_text(value)
    return normalized


def sample_rows(kind):
    today = date.today()
    patients = [
        {
            "VIP": "ΣΜΧΟΣ",
            "PCD": "ΙΠΤ",
            "CHR": "ΝΙΚΟΛΑΟΣ",
            "NAME": "ΠΑΠΑΔΟΠΟΥΛΟΣ",
            "UNIT_NAME": "251 ΓΝΑ",
            "STREET": "ΜΕΣΟΓΕΙΩΝ 1",
            "CITY": "ΑΘΗΝΑ",
            "CFE": "ΣΤΡΑΤΙΩΤΙΚΟΣ",
            "CFE_TYPE01": "ΣΤΡ.ΠΡΣ",
            "HM_EIS": (today - timedelta(days=2)).isoformat(),
            "WDS": "3",
            "DEP": "ΠΑΘΟΛΟΓΙΚΗ",
            "ROOM": "305",
            "PAT": "900001",
            "PERSCODE": "ΑΒ123456",
            "ZIP": "11525",
            "PHONE": "2100000000",
            "AMKA_DOC": "01017000000",
        },
        {
            "VIP": "ΙΔΙΩΤΗΣ",
            "PCD": "",
            "CHR": "ΜΑΡΙΑ",
            "NAME": "ΓΕΩΡΓΙΟΥ",
            "UNIT_NAME": "",
            "STREET": "ΚΗΦΙΣΙΑΣ 10",
            "CITY": "ΧΑΛΑΝΔΡΙ",
            "CFE": "ΛΟΙΠΟΙ",
            "CFE_TYPE01": "ΛΟΙΠΟΙ",
            "HM_EIS": (today - timedelta(days=1)).isoformat(),
            "WDS": "2",
            "DEP": "ΧΕΙΡΟΥΡΓΙΚΗ",
            "ROOM": "214",
            "PAT": "900002",
            "PERSCODE": "ΑΖ987654",
            "ZIP": "15231",
            "PHONE": "2101111111",
            "AMKA_DOC": "02028000000",
        },
    ]
    if kind in {"patients", "movement"}:
        return patients
    return [
        {
            "PAT": "900003",
            "NAMECHR": "ΔΗΜΗΤΡΙΟΥ ΚΩΝΣΤΑΝΤΙΝΟΣ",
            "DISDCALC": (today - timedelta(days=1)).isoformat(),
            "DEP": "ΟΡΘΟΠΕΔΙΚΗ",
            "NAME": "ΔΗΜΟΣΙΟ",
            "VIP": "ΣΓΟΣ",
            "KCODE": "K123",
        }
    ]


def init_oracle_client_once():
    global _oracle_client_initialized
    if _oracle_client_initialized:
        return
    if not ORACLE_CLIENT_LIB_DIR:
        _oracle_client_initialized = True
        return
    import oracledb

    kwargs = {"lib_dir": ORACLE_CLIENT_LIB_DIR}
    if ORACLE_CONFIG_DIR:
        kwargs["config_dir"] = ORACLE_CONFIG_DIR
    oracledb.init_oracle_client(**kwargs)
    _oracle_client_initialized = True


def fetch_medico_rows(query, kind):
    if MEDICO_SAMPLE_MODE:
        return [normalize_external_row(row) for row in sample_rows(kind)]
    user = os.getenv("MEDICO_USER")
    password = os.getenv("MEDICO_PASSWORD")
    dsn = os.getenv("MEDICO_DSN", "MEDICOSRV")
    if not user or not password:
        raise RuntimeError("MEDICO_USER και MEDICO_PASSWORD δεν έχουν οριστεί.")
    init_oracle_client_once()
    import oracledb

    if ORACLE_CONFIG_DIR:
        os.environ.setdefault("TNS_ADMIN", ORACLE_CONFIG_DIR)
        if hasattr(oracledb, "defaults") and hasattr(oracledb.defaults, "config_dir"):
            oracledb.defaults.config_dir = ORACLE_CONFIG_DIR

    with oracledb.connect(user=user, password=password, dsn=dsn) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query)
            columns = [col[0] for col in cursor.description]
            rows = []
            for db_row in cursor.fetchall():
                rows.append(normalize_external_row(dict(zip(columns, db_row))))
            return rows


def patient_hash_from_row(row):
    payload = {key: row.get(key, "") for key in PATIENT_MEDICO_MAP}
    encoded = json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def apply_medico_row(patient, row):
    changed = False
    for source_key, attr in PATIENT_MEDICO_MAP.items():
        value = row.get(source_key)
        if attr == "hm_eis":
            value = parse_date(value)
        else:
            value = as_text(value) or None
        if getattr(patient, attr) != value:
            setattr(patient, attr, value)
            changed = True
    raw_json = json.dumps(row, ensure_ascii=False, sort_keys=True)
    row_hash = patient_hash_from_row(row)
    if patient.raw_medico_json != raw_json:
        patient.raw_medico_json = raw_json
        changed = True
    if patient.medico_hash != row_hash:
        patient.medico_hash = row_hash
        changed = True
    patient.last_seen_at = utcnow()
    patient.last_synced_at = utcnow()
    return changed


def sync_patients_from_medico(actor="system"):
    sync_log = SyncLog(actor=actor)
    db.session.add(sync_log)
    db.session.flush()
    try:
        rows = fetch_medico_rows(PATIENT_QUERY, "patients")
        sync_log.rows_seen = len(rows)
        seen_pats = set()
        for row in rows:
            pat = as_text(row.get("PAT"))
            if not pat:
                continue
            seen_pats.add(pat)
            patient = PatientRecord.query.filter_by(pat=pat).first()
            if not patient:
                patient = PatientRecord(pat=pat)
                db.session.add(patient)
                sync_log.inserted += 1
            elif apply_medico_row(patient, row):
                sync_log.updated += 1
            if patient.id is None or not patient.raw_medico_json:
                apply_medico_row(patient, row)

        if MEDICO_DELETE_MISSING:
            existing_count = PatientRecord.query.count()
            if rows or not (MEDICO_SKIP_DELETE_ON_EMPTY and existing_count > 0):
                stale = PatientRecord.query.filter(~PatientRecord.pat.in_(seen_pats)).all() if seen_pats else PatientRecord.query.all()
                for patient in stale:
                    log_action("PATIENT_REMOVED_BY_SYNC", "PatientRecord", patient.pat, patient.full_name)
                    db.session.delete(patient)
                    sync_log.deleted += 1
            elif existing_count > 0:
                sync_log.message = "Το MEDICO επέστρεψε μηδενικές γραμμές. Παραλείφθηκαν διαγραφές για προστασία από προσωρινή αστοχία."

        sync_log.status = "OK"
        sync_log.ended_at = utcnow()
        log_action(
            "MEDICO_SYNC",
            "SyncLog",
            sync_log.id,
            f"rows={sync_log.rows_seen}, inserted={sync_log.inserted}, updated={sync_log.updated}, deleted={sync_log.deleted}",
        )
        db.session.commit()
        return sync_log
    except Exception as exc:
        db.session.rollback()
        sync_log = SyncLog(
            actor=actor,
            ended_at=utcnow(),
            status="FAILED",
            message=str(exc)[:1000],
        )
        db.session.add(sync_log)
        log_action("MEDICO_SYNC_FAILED", "SyncLog", None, str(exc)[:1000])
        db.session.commit()
        return sync_log


def filter_rows(rows, query):
    if not query:
        return rows
    needle = query.strip().lower()
    return [row for row in rows if any(needle in as_text(value).lower() for value in row.values())]


def make_workbook(title, columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_fill = PatternFill("solid", fgColor="E8F1FB")
    header_font = Font(bold=True, color="172033")
    for col_index, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_index, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row_index, row in enumerate(rows, start=2):
        for col_index, (key, _) in enumerate(columns, start=1):
            value = getattr(row, key, None) if not isinstance(row, dict) else row.get(key)
            if isinstance(value, (datetime, date)):
                value = value.strftime("%d/%m/%Y")
            ws.cell(row=row_index, column=col_index, value=value)
    for col in ws.columns:
        max_len = max(len(as_text(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 36)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def qr_png_response(value):
    image = qrcode.make(value)
    output = io.BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return send_file(output, mimetype="image/png", download_name="qr.png")


def save_authorization_file(file_storage, visitor):
    if not file_storage or not file_storage.filename:
        return
    folder = os.path.join(UPLOAD_FOLDER, "visitors")
    os.makedirs(folder, exist_ok=True)
    original = secure_filename(file_storage.filename)
    extension = os.path.splitext(original)[1].lower()
    filename = f"{visitor.qr_token}{extension}"
    path = os.path.join(folder, filename)
    file_storage.save(path)
    visitor.authorization_filename = filename
    visitor.authorization_original_filename = original


def seed_local_users():
    seeds = [
        ("ADMIN_USERNAME", "ADMIN_PASSWORD", "admin", "admin"),
        ("SECURITY_OFFICE_USERNAME", "SECURITY_OFFICE_PASSWORD", "security_office", "security_office"),
        ("TRANSIT_OFFICE_USERNAME", "TRANSIT_OFFICE_PASSWORD", "transit_office", "transit_office"),
        ("SECURITY_POINT_USERNAME", "SECURITY_POINT_PASSWORD", "security_point", "security_point"),
    ]
    changed = False
    for user_env, pass_env, default_username, role in seeds:
        username = os.getenv(user_env, default_username).strip()
        password = os.getenv(pass_env, "").strip()
        if not username or not password:
            continue
        user = User.query.filter_by(username=username).first()
        if not user:
            user = User(username=username, password_hash=generate_password_hash(password), role=role)
            db.session.add(user)
            changed = True
        elif user.role != role and username != os.getenv("ADMIN_USERNAME", "admin"):
            user.role = role
            changed = True
    if changed:
        db.session.commit()


@app.errorhandler(403)
def forbidden(_error):
    return render_template("error.html", status_code=403, message="Δεν έχετε δικαίωμα πρόσβασης σε αυτή τη λειτουργία."), 403


@app.errorhandler(CSRFError)
def csrf_error(error):
    flash("Η φόρμα έληξε. Δοκιμάστε ξανά.", "warning")
    return redirect(request.referrer or url_for("index"))


@app.route("/health")
def health():
    return {
        "ok": True,
        "app": "SEC_OFFICE",
        "database": "connected",
        "time": utcnow().isoformat(),
    }


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and user.is_active_local and check_password_hash(user.password_hash, password):
            login_user(user)
            session.permanent = True
            log_action("LOGIN", "User", user.username, f"Local login as {role_label(user.role)}")
            db.session.commit()
            return redirect(request.args.get("next") or url_for("index"))
        flash("Λάθος username ή κωδικός.", "danger")
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    log_action("LOGOUT", "User", current_actor())
    db.session.commit()
    session.clear()
    logout_user()
    return redirect(url_for("login"))


@app.route("/change_password", methods=["GET", "POST"])
@login_required
def change_password():
    if getattr(current_user, "auth_method", "") == "sso":
        flash("Η αλλαγή κωδικού γίνεται στο κεντρικό σύστημα auth.", "info")
        return redirect(url_for("index"))
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")
        if not check_password_hash(current_user.password_hash, current_password):
            flash("Ο τρέχων κωδικός δεν είναι σωστός.", "danger")
        elif len(new_password) < 8:
            flash("Ο νέος κωδικός πρέπει να έχει τουλάχιστον 8 χαρακτήρες.", "warning")
        elif new_password != confirm_password:
            flash("Η επιβεβαίωση δεν ταιριάζει.", "warning")
        else:
            current_user.password_hash = generate_password_hash(new_password)
            log_action("PASSWORD_CHANGE", "User", current_user.username)
            db.session.commit()
            flash("Ο κωδικός άλλαξε.", "success")
            return redirect(url_for("index"))
    return render_template("change_password.html")


@app.route("/")
@login_required
def index():
    latest_sync = SyncLog.query.order_by(SyncLog.started_at.desc()).first()
    recent_audit = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(8).all()
    return render_template(
        "index.html",
        patient_count=PatientRecord.query.count(),
        visitor_count=VisitorRecord.query.filter_by(status="active").count(),
        valid_visitor_count=sum(1 for visitor in VisitorRecord.query.filter_by(status="active").all() if visitor.is_currently_valid),
        latest_sync=latest_sync,
        recent_audit=recent_audit,
    )


@app.route("/patients")
@login_required
def patients():
    require_permission(can_view_patients)
    q = request.args.get("q", "").strip()
    dep = request.args.get("dep", "").strip()
    query = PatientRecord.query
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                PatientRecord.pat.ilike(like),
                PatientRecord.chr.ilike(like),
                PatientRecord.name.ilike(like),
                PatientRecord.room.ilike(like),
                PatientRecord.escort_one.ilike(like),
                PatientRecord.escort_two.ilike(like),
            )
        )
    if dep:
        query = query.filter(PatientRecord.dep == dep)
    records = query.order_by(PatientRecord.dep.asc(), PatientRecord.room.asc(), PatientRecord.name.asc()).all()
    deps = [row[0] for row in db.session.query(PatientRecord.dep).filter(PatientRecord.dep.isnot(None)).distinct().order_by(PatientRecord.dep).all()]
    latest_sync = SyncLog.query.order_by(SyncLog.started_at.desc()).first()
    return render_template("patients.html", records=records, deps=deps, q=q, dep=dep, latest_sync=latest_sync)


@app.route("/patients/sync", methods=["POST"])
@login_required
def patients_sync():
    require_permission(can_sync_patients)
    sync_log = sync_patients_from_medico(actor=current_actor())
    if sync_log.status == "OK":
        flash("Ο συγχρονισμός MEDICO ολοκληρώθηκε.", "success")
    else:
        flash(f"Ο συγχρονισμός απέτυχε: {sync_log.message}", "danger")
    return redirect(url_for("patients"))


@app.route("/patients/export.xlsx")
@login_required
def patients_export():
    require_permission(can_view_patients)
    rows = PatientRecord.query.order_by(PatientRecord.dep.asc(), PatientRecord.room.asc(), PatientRecord.name.asc()).all()
    columns = PATIENT_COLUMNS + [("escort_one", "ΣΥΝΟΔΟΣ 1"), ("escort_two", "ΣΥΝΟΔΟΣ 2"), ("security_notes", "ΣΗΜΕΙΩΣΕΙΣ")]
    output = make_workbook("Νοσηλευόμενοι", columns, rows)
    log_action("EXPORT_PATIENTS", "PatientRecord", "xlsx", f"{len(rows)} rows")
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"sec-office-patients-{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/patients/<int:patient_id>", methods=["GET", "POST"])
@login_required
def patient_detail(patient_id):
    patient = PatientRecord.query.get_or_404(patient_id)
    if request.method == "POST":
        require_permission(can_edit_patient_access)
        patient.escort_one = request.form.get("escort_one", "").strip() or None
        patient.escort_two = request.form.get("escort_two", "").strip() or None
        patient.security_notes = request.form.get("security_notes", "").strip() or None
        log_action("PATIENT_MANUAL_UPDATE", "PatientRecord", patient.pat, patient.full_name)
        db.session.commit()
        flash("Τα χειροκίνητα στοιχεία αποθηκεύτηκαν.", "success")
        return redirect(url_for("patient_detail", patient_id=patient.id))
    require_permission(can_view_patients)
    return render_template("patient_detail.html", patient=patient)


@app.route("/patients/<int:patient_id>/qr.png")
@login_required
def patient_qr(patient_id):
    patient = PatientRecord.query.get_or_404(patient_id)
    require_permission(can_view_patients)
    return qr_png_response(url_for("scan_token", token=patient.qr_token, _external=True))


@app.route("/patients/<int:patient_id>/label")
@login_required
def patient_label(patient_id):
    patient = PatientRecord.query.get_or_404(patient_id)
    require_permission(can_view_patients)
    return render_template("label.html", kind="patient", record=patient)


@app.route("/visitors")
@login_required
def visitors():
    require_permission(can_view_visitors)
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "active").strip()
    query = VisitorRecord.query
    if status:
        query = query.filter_by(status=status)
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                VisitorRecord.full_name.ilike(like),
                VisitorRecord.identity_number.ilike(like),
                VisitorRecord.organization.ilike(like),
                VisitorRecord.destination.ilike(like),
                VisitorRecord.authorization_doc_no.ilike(like),
            )
        )
    records = query.order_by(VisitorRecord.updated_at.desc()).all()
    return render_template("visitors.html", records=records, q=q, status=status)


@app.route("/visitors/new", methods=["GET", "POST"])
@login_required
def visitor_new():
    require_permission(can_manage_visitors)
    visitor = VisitorRecord()
    if request.method == "POST":
        fill_visitor_from_form(visitor)
        visitor.created_by = current_actor()
        visitor.updated_by = current_actor()
        db.session.add(visitor)
        db.session.flush()
        save_authorization_file(request.files.get("authorization_file"), visitor)
        log_action("VISITOR_CREATE", "VisitorRecord", visitor.qr_token, visitor.full_name)
        db.session.commit()
        flash("Η εγγραφή επισκέπτη δημιουργήθηκε.", "success")
        return redirect(url_for("visitors"))
    return render_template("visitor_form.html", visitor=visitor, mode="new")


@app.route("/visitors/<int:visitor_id>", methods=["GET", "POST"])
@login_required
def visitor_edit(visitor_id):
    visitor = VisitorRecord.query.get_or_404(visitor_id)
    if request.method == "POST":
        require_permission(can_manage_visitors)
        fill_visitor_from_form(visitor)
        visitor.updated_by = current_actor()
        save_authorization_file(request.files.get("authorization_file"), visitor)
        log_action("VISITOR_UPDATE", "VisitorRecord", visitor.qr_token, visitor.full_name)
        db.session.commit()
        flash("Η εγγραφή επισκέπτη αποθηκεύτηκε.", "success")
        return redirect(url_for("visitors"))
    require_permission(can_view_visitors)
    return render_template("visitor_form.html", visitor=visitor, mode="edit")


def fill_visitor_from_form(visitor):
    visitor.full_name = request.form.get("full_name", "").strip()
    visitor.identity_number = request.form.get("identity_number", "").strip() or None
    visitor.phone = request.form.get("phone", "").strip() or None
    visitor.organization = request.form.get("organization", "").strip() or None
    visitor.role_title = request.form.get("role_title", "").strip() or None
    visitor.visit_reason = request.form.get("visit_reason", "").strip() or None
    visitor.destination = request.form.get("destination", "").strip() or None
    visitor.access_scope = request.form.get("access_scope", "").strip() or None
    visitor.authorization_doc_no = request.form.get("authorization_doc_no", "").strip() or None
    visitor.authorization_issuer = request.form.get("authorization_issuer", "").strip() or None
    visitor.authorization_date = parse_date(request.form.get("authorization_date", ""))
    visitor.valid_from = parse_datetime_local(request.form.get("valid_from", ""))
    visitor.valid_until = parse_datetime_local(request.form.get("valid_until", ""))
    visitor.status = request.form.get("status", "active").strip() or "active"
    visitor.notes = request.form.get("notes", "").strip() or None


@app.route("/visitors/<int:visitor_id>/qr.png")
@login_required
def visitor_qr(visitor_id):
    visitor = VisitorRecord.query.get_or_404(visitor_id)
    require_permission(can_view_visitors)
    return qr_png_response(url_for("scan_token", token=visitor.qr_token, _external=True))


@app.route("/visitors/<int:visitor_id>/label")
@login_required
def visitor_label(visitor_id):
    visitor = VisitorRecord.query.get_or_404(visitor_id)
    require_permission(can_view_visitors)
    return render_template("label.html", kind="visitor", record=visitor)


@app.route("/visitors/<int:visitor_id>/authorization")
@login_required
def visitor_authorization(visitor_id):
    visitor = VisitorRecord.query.get_or_404(visitor_id)
    require_permission(can_view_visitors)
    if not visitor.authorization_filename:
        abort(404)
    path = os.path.join(UPLOAD_FOLDER, "visitors", visitor.authorization_filename)
    return send_file(path, as_attachment=True, download_name=visitor.authorization_original_filename or visitor.authorization_filename)


@app.route("/visitors/export.xlsx")
@login_required
def visitors_export():
    require_permission(can_view_visitors)
    rows = VisitorRecord.query.order_by(VisitorRecord.updated_at.desc()).all()
    columns = [
        ("full_name", "ΟΝΟΜΑΤΕΠΩΝΥΜΟ"),
        ("identity_number", "ΤΑΥΤΟΤΗΤΑ"),
        ("phone", "ΤΗΛ."),
        ("organization", "ΦΟΡΕΑΣ"),
        ("visit_reason", "ΛΟΓΟΣ"),
        ("destination", "ΠΡΟΟΡΙΣΜΟΣ"),
        ("authorization_doc_no", "ΕΓΓΡΑΦΟ"),
        ("valid_from", "ΙΣΧΥΕΙ ΑΠΟ"),
        ("valid_until", "ΙΣΧΥΕΙ ΕΩΣ"),
        ("status", "STATUS"),
        ("notes", "ΣΗΜΕΙΩΣΕΙΣ"),
    ]
    output = make_workbook("Επισκέπτες", columns, rows)
    log_action("EXPORT_VISITORS", "VisitorRecord", "xlsx", f"{len(rows)} rows")
    db.session.commit()
    return send_file(
        output,
        as_attachment=True,
        download_name=f"sec-office-visitors-{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/scan", methods=["GET", "POST"])
@login_required
def scan():
    require_permission(can_scan_qr)
    if request.method == "POST":
        token = request.form.get("token", "").strip()
        if not token:
            flash("Δεν διαβάστηκε QR.", "warning")
            return redirect(url_for("scan"))
        return redirect(url_for("scan_token", token=token))
    return render_template("scan.html")


@app.route("/scan/<path:token>")
@login_required
def scan_token(token):
    require_permission(can_scan_qr)
    cleaned = token.strip()
    if "/scan/" in cleaned:
        cleaned = cleaned.rsplit("/scan/", 1)[-1].strip()
    patient = PatientRecord.query.filter_by(qr_token=cleaned).first()
    visitor = VisitorRecord.query.filter_by(qr_token=cleaned).first()
    if patient:
        patient.last_scanned_at = utcnow()
        patient.last_scanned_by = current_actor()
        log_action("SCAN_PATIENT_QR", "PatientRecord", patient.pat, patient.full_name)
        db.session.commit()
        return render_template("scan_result.html", kind="patient", record=patient)
    if visitor:
        visitor.last_scanned_at = utcnow()
        visitor.last_scanned_by = current_actor()
        log_action("SCAN_VISITOR_QR", "VisitorRecord", visitor.qr_token, visitor.full_name)
        db.session.commit()
        return render_template("scan_result.html", kind="visitor", record=visitor)
    log_action("SCAN_UNKNOWN_QR", "QR", cleaned[:120])
    db.session.commit()
    flash("Το QR δεν βρέθηκε στην εφαρμογή.", "danger")
    return redirect(url_for("scan"))


@app.route("/transit/patients")
@login_required
def transit_patients():
    require_permission(can_view_transit)
    q = request.args.get("q", "").strip()
    error = None
    rows = []
    try:
        rows = filter_rows(fetch_medico_rows(MOVEMENT_QUERY, "movement"), q)
    except Exception as exc:
        error = str(exc)
    return render_template(
        "external_table.html",
        title="Στοιχεία ασθενών για το Γρ. Διακίνησης",
        endpoint="transit_patients",
        columns=MOVEMENT_COLUMNS,
        rows=rows,
        q=q,
        error=error,
        export_endpoint="transit_patients_export",
    )


@app.route("/transit/patients/export.xlsx")
@login_required
def transit_patients_export():
    require_permission(can_view_transit)
    rows = fetch_medico_rows(MOVEMENT_QUERY, "movement")
    output = make_workbook("Διακίνηση", MOVEMENT_COLUMNS, rows)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"sec-office-transit-{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/transit/no-admin-discharge")
@login_required
def no_admin_discharge():
    require_permission(can_view_transit)
    q = request.args.get("q", "").strip()
    error = None
    rows = []
    try:
        rows = filter_rows(fetch_medico_rows(NO_EXIT_QUERY, "no_exit"), q)
    except Exception as exc:
        error = str(exc)
    return render_template(
        "external_table.html",
        title="Ασθενείς χωρίς διοικητικό εξιτήριο",
        endpoint="no_admin_discharge",
        columns=NO_EXIT_COLUMNS,
        rows=rows,
        q=q,
        error=error,
        export_endpoint="no_admin_discharge_export",
    )


@app.route("/transit/no-admin-discharge/export.xlsx")
@login_required
def no_admin_discharge_export():
    require_permission(can_view_transit)
    rows = fetch_medico_rows(NO_EXIT_QUERY, "no_exit")
    output = make_workbook("Χωρίς Εξιτήριο", NO_EXIT_COLUMNS, rows)
    return send_file(
        output,
        as_attachment=True,
        download_name=f"sec-office-no-admin-discharge-{date.today().isoformat()}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/audit")
@login_required
def audit():
    require_permission(can_view_audit)
    q = request.args.get("q", "").strip()
    role = request.args.get("role", "").strip()
    action = request.args.get("action", "").strip()
    query = AuditLog.query
    if role:
        query = query.filter(AuditLog.role == role)
    if action:
        query = query.filter(AuditLog.action.ilike(f"%{action}%"))
    if q:
        like = f"%{q}%"
        query = query.filter(
            db.or_(
                AuditLog.actor.ilike(like),
                AuditLog.target_id.ilike(like),
                AuditLog.details.ilike(like),
                AuditLog.action.ilike(like),
            )
        )
    rows = query.order_by(AuditLog.created_at.desc()).limit(500).all()
    return render_template("audit.html", rows=rows, q=q, role=role, action=action)


@app.route("/manage_users")
@login_required
def manage_users():
    require_permission(can_manage_users)
    users = User.query.order_by(User.username.asc()).all()
    central_auth = {
        "realm": CENTRAL_AUTH_REALM,
        "admin_url": CENTRAL_AUTH_ADMIN_URL,
        "users_url": CENTRAL_AUTH_USERS_URL,
        "groups_url": CENTRAL_AUTH_GROUPS_URL,
        "groups": [
            (SSO_APP_USER_GROUP, "Βασική πρόσβαση εφαρμογής"),
            (SSO_APP_ADMIN_GROUP, "Admin εφαρμογής"),
            (SSO_SECURITY_OFFICE_GROUP, "Γραφείο Ασφαλείας"),
            (SSO_TRANSIT_OFFICE_GROUP, "Γραφείο Διακίνησης"),
            (SSO_SECURITY_POINT_GROUP, "Security point"),
            (SSO_GLOBAL_ADMIN_GROUP, "Global admins"),
        ],
    }
    return render_template("manage_users.html", users=users, central_auth=central_auth)


@app.route("/users/create", methods=["POST"])
@login_required
def create_user():
    require_permission(can_manage_users)
    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", "security_point")
    if role not in ROLES:
        role = "security_point"
    if not username or len(password) < 8:
        flash("Συμπληρώστε username και κωδικό τουλάχιστον 8 χαρακτήρων.", "warning")
        return redirect(url_for("manage_users"))
    user = User(username=username, password_hash=generate_password_hash(password), role=role)
    db.session.add(user)
    try:
        log_action("USER_CREATE", "User", username, role)
        db.session.commit()
        flash("Ο χρήστης δημιουργήθηκε.", "success")
    except IntegrityError:
        db.session.rollback()
        flash("Υπάρχει ήδη χρήστης με αυτό το username.", "danger")
    return redirect(url_for("manage_users"))


@app.route("/users/<int:user_id>/update", methods=["POST"])
@login_required
def update_user(user_id):
    require_permission(can_manage_users)
    user = User.query.get_or_404(user_id)
    role = request.form.get("role", user.role)
    if role in ROLES:
        user.role = role
    user.is_active_local = bool(request.form.get("is_active_local"))
    new_password = request.form.get("new_password", "")
    if new_password:
        if len(new_password) < 8:
            flash("Ο νέος κωδικός πρέπει να έχει τουλάχιστον 8 χαρακτήρες.", "warning")
            return redirect(url_for("manage_users"))
        user.password_hash = generate_password_hash(new_password)
    log_action("USER_UPDATE", "User", user.username, user.role)
    db.session.commit()
    flash("Ο χρήστης ενημερώθηκε.", "success")
    return redirect(url_for("manage_users"))


@app.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
def delete_user(user_id):
    require_permission(can_manage_users)
    user = User.query.get_or_404(user_id)
    if user.username == os.getenv("ADMIN_USERNAME", "admin"):
        flash("Ο βασικός admin δεν διαγράφεται.", "warning")
        return redirect(url_for("manage_users"))
    log_action("USER_DELETE", "User", user.username)
    db.session.delete(user)
    db.session.commit()
    flash("Ο χρήστης διαγράφηκε.", "success")
    return redirect(url_for("manage_users"))


def get_backup_password():
    return (os.getenv("BACKUP_PASSWORD") or "change-me-local-backup-password").encode("utf-8")


def safe_backup_path(filename):
    backup_root = os.path.abspath(BACKUP_FOLDER)
    path = os.path.abspath(os.path.join(backup_root, filename))
    if not path.startswith(backup_root):
        raise ValueError("Invalid backup path")
    return path


def database_uri_path():
    uri = app.config["SQLALCHEMY_DATABASE_URI"]
    if uri.startswith("sqlite:///"):
        return uri.replace("sqlite:///", "", 1)
    return None


def create_backup_archive(backup_type="manual", actor=None):
    timestamp = utcnow().strftime("%Y%m%d-%H%M%S")
    filename = f"sec-office-backup-{timestamp}-{backup_type}.zip"
    backup_path = safe_backup_path(filename)
    with tempfile.TemporaryDirectory() as tmpdir:
        manifest = {
            "app": "SEC_OFFICE",
            "created_at": utcnow().isoformat(),
            "backup_type": backup_type,
            "actor": actor,
            "database_uri_kind": "sqlite" if database_uri_path() else "external",
        }
        if BACKUP_INCLUDE_DATABASE:
            sqlite_path = database_uri_path()
            if sqlite_path and os.path.exists(sqlite_path):
                db.session.remove()
                shutil.copy2(sqlite_path, os.path.join(tmpdir, "sec_office.sqlite3"))
                manifest["database"] = "sec_office.sqlite3"
            else:
                dump_path = os.path.join(tmpdir, "database.dump")
                database_url = app.config["SQLALCHEMY_DATABASE_URI"]
                subprocess.run(["pg_dump", database_url, "-f", dump_path], check=True, timeout=120)
                manifest["database"] = "database.dump"
        uploads_path = os.path.join(tmpdir, "uploads")
        if os.path.isdir(UPLOAD_FOLDER):
            shutil.copytree(UPLOAD_FOLDER, uploads_path, dirs_exist_ok=True)
            manifest["uploads"] = "uploads"
        with open(os.path.join(tmpdir, "manifest.json"), "w", encoding="utf-8") as handle:
            json.dump(manifest, handle, ensure_ascii=False, indent=2)
        with pyzipper.AESZipFile(backup_path, "w", compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
            zf.setpassword(get_backup_password())
            for root, _dirs, files in os.walk(tmpdir):
                for file_name in files:
                    full_path = os.path.join(root, file_name)
                    arcname = os.path.relpath(full_path, tmpdir)
                    zf.write(full_path, arcname)
    sha = hashlib.sha256()
    with open(backup_path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            sha.update(chunk)
    return backup_path, filename, os.path.getsize(backup_path), sha.hexdigest()


def create_backup_record(backup_type="manual", actor=None):
    path, filename, size, sha = create_backup_archive(backup_type=backup_type, actor=actor)
    record = BackupRecord(
        created_by=actor,
        auth_method=getattr(current_user, "auth_method", None) if current_user.is_authenticated else "system",
        backup_type=backup_type,
        filename=filename,
        path=path,
        size_bytes=size,
        sha256=sha,
    )
    db.session.add(record)
    log_action("BACKUP_CREATE", "BackupRecord", filename)
    db.session.commit()
    prune_backups()
    return record


def verify_backup(record):
    if not os.path.exists(record.path):
        record.status = "VERIFY_FAILED"
        record.verify_message = "Το αρχείο δεν υπάρχει."
        record.verified_at = utcnow()
        return False
    sha = hashlib.sha256()
    with open(record.path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            sha.update(chunk)
    if sha.hexdigest() != record.sha256:
        record.status = "VERIFY_FAILED"
        record.verify_message = "SHA-256 mismatch."
        record.verified_at = utcnow()
        return False
    try:
        with pyzipper.AESZipFile(record.path) as zf:
            zf.setpassword(get_backup_password())
            bad_file = zf.testzip()
        if bad_file:
            record.status = "VERIFY_FAILED"
            record.verify_message = f"Zip test failed: {bad_file}"
            return False
        record.status = "VERIFIED"
        record.verify_message = "OK"
        return True
    finally:
        record.verified_at = utcnow()


def prune_backups():
    if BACKUP_RETENTION_COUNT <= 0:
        return
    records = BackupRecord.query.order_by(BackupRecord.created_at.desc()).all()
    for record in records[BACKUP_RETENTION_COUNT:]:
        if record.status == "PRUNED":
            continue
        try:
            if os.path.exists(record.path):
                os.remove(record.path)
            record.status = "PRUNED"
            record.verify_message = f"Pruned by retention policy: keep last {BACKUP_RETENTION_COUNT}"
        except OSError as exc:
            record.verify_message = str(exc)
    db.session.commit()


@app.route("/manage_backups")
@login_required
def manage_backups():
    require_permission(can_manage_backups)
    backups = BackupRecord.query.order_by(BackupRecord.created_at.desc()).limit(100).all()
    return render_template(
        "manage_backups.html",
        backups=backups,
        retention_count=BACKUP_RETENTION_COUNT,
        include_database=BACKUP_INCLUDE_DATABASE,
    )


@app.route("/backups/create", methods=["POST"])
@login_required
def create_backup_route():
    require_permission(can_manage_backups)
    try:
        record = create_backup_record("manual", current_actor())
        flash(f"Δημιουργήθηκε backup: {record.filename}", "success")
    except Exception as exc:
        db.session.rollback()
        flash(f"Αποτυχία backup: {exc}", "danger")
    return redirect(url_for("manage_backups"))


@app.route("/backups/<int:backup_id>/verify", methods=["POST"])
@login_required
def verify_backup_route(backup_id):
    require_permission(can_manage_backups)
    record = BackupRecord.query.get_or_404(backup_id)
    verify_backup(record)
    log_action("BACKUP_VERIFY", "BackupRecord", record.filename, record.verify_message)
    db.session.commit()
    flash(record.verify_message or record.status, "success" if record.status == "VERIFIED" else "danger")
    return redirect(url_for("manage_backups"))


@app.route("/backups/<int:backup_id>/download")
@login_required
def download_backup_route(backup_id):
    require_permission(can_manage_backups)
    record = BackupRecord.query.get_or_404(backup_id)
    if record.status == "PRUNED" or not os.path.exists(record.path):
        abort(404)
    log_action("BACKUP_DOWNLOAD", "BackupRecord", record.filename)
    db.session.commit()
    return send_file(record.path, as_attachment=True, download_name=record.filename)


def scheduled_sync_job():
    with app.app_context():
        sync_patients_from_medico(actor="scheduler")


def scheduled_backup_job():
    with app.app_context():
        try:
            create_backup_record("scheduled", "scheduler")
        except Exception:
            db.session.rollback()


def start_scheduler_once():
    if bool_env("DISABLE_SCHEDULER", "0"):
        return None
    if app.debug and os.getenv("WERKZEUG_RUN_MAIN") != "true":
        return None
    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(func=scheduled_sync_job, trigger="interval", minutes=MEDICO_SYNC_INTERVAL_MINUTES, id="medico-sync", replace_existing=True)
    scheduler.add_job(func=scheduled_backup_job, trigger="cron", hour=3, minute=10, id="daily-backup", replace_existing=True)
    scheduler.start()
    return scheduler


def create_all_with_retry(attempts=10, delay_seconds=2):
    for attempt in range(1, attempts + 1):
        try:
            db.create_all()
            return
        except Exception:
            db.session.rollback()
            if attempt == attempts:
                raise
            time.sleep(delay_seconds)


@app.route("/robots.txt")
def robots():
    return Response("User-agent: *\nDisallow: /\n", mimetype="text/plain")


with app.app_context():
    create_all_with_retry()
    seed_local_users()

scheduler = start_scheduler_once()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5055")), debug=bool_env("FLASK_DEBUG", "1"))
